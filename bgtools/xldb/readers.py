import os
import sys

import xlrd

import openpyxl as pyxl

from urllib2 import urlopen, URLError
from datetime import date, timedelta
from tempfile import mkstemp

from .utils import *

class XLSReader(object):
    def __init__(self, url, localfile=True, hash_comments=True):
        self.book = None
        self.hash_comments = hash_comments
        
        if localfile:
            try:
                self.book = xlrd.open_workbook(on_demand=True,
                                               filename=url)
            except:
                print("Error on %s" % url)
        else:
            try:
                conn = urlopen(url)
                
            except URLError as strerr:
                print("\nURL Error reading url: %s\n %s" % (url, strerr))
            
            except: 
                print("\nGeneral Error reading url: %s\n" % url)
                
            else:
                try:
                    self.book = xlrd.open_workbook(on_demand=True,
                                                   file_contents=conn.read())
                except:
                    print("Error on %s" % url)
            
            finally:
                conn.close()
            
        if self.book:
            self.datemode = self.book.datemode
                
    def xlCellValue(self, x):
        return xlValue(x, self.datemode, self.hash_comments)

    def sheet(self, sheet=None):
        '''
        Returns worksheet object by name or index, 
        
        Example:
        > wb = XLSReader('test.xls')
        > wb.sheet('sheet2')   # returns sheet named 'sheet2'
        > wb.sheet(3)          # returns sheet with index 3
        
        Default (if no argument is pased) returns sheet index = 0 
        
        '''
        if type(sheet) == int:
            sheet_name, sheet_index = None, sheet
        else:
            sheet_name, sheet_index = sheet, 0
        
        try:
            if sheet_name:
                self.sh = self.book.sheet_by_name(sheet)
            else:
                self.sh = self.book.sheet_by_index(sheet)
        except:
            print("Invalid sheet %s %s" % (sheet, type(sheet)))
            return None
        
        self.ncolumns = self.sh.ncols
        self.nrows = self.sh.nrows
        
        return self.sh
        
    def sheet_as_db(self, sheet=None, 
                    header=True, dkey=0,
                    startrow=0, numrows=None):
        '''
        Reads rows in a given sheet. Returns (keys, hdr, qdata)
        
        sheet:  Name of sheet or 0-indexed sheet number.          
                        
        header: True--return rows as dict objects, with 'startrow' as keys.
                False--return rows as list
        
        dkey:   Column to serve as the dict key for qdata.
                '-1' uses the row number as key.
        
        startrow: Begin reading at this row (0 indexed), ignore ealier rows                
        
        numrows: number of rows to read, None=read all
        
        '''
        class sheetdb(object):
            pass
            
        self.sh = self.sheet(sheet)
        
        cleanrow_ = lambda row_: [x if x is not '' else None for x in row_]
            
        startloc = 1 if dkey == 0 else 0
        
        hdr = None
        if header:
            hdr = [get_xlvalue(h) for h in self.sh.row(startrow)]
            startrow += 1
            
            def rowValues(row_, loc): 
                return dict(zip(hdr[loc:],
                                cleanrow_(row_[loc:])))
            
                                                   
        else:
            rowValues = lambda row_, loc: cleanrow_(row_[loc:])
        
        qdata = []
        key_column = []
        
        numrows = self.sh.nrows if not numrows else startrow + numrows
            
        for xrow in range(startrow, numrows):
            try:
                xr = map(self.xlCellValue, self.sh.row(xrow))
                
            except:
                #skips a row if there's a problem
                print("problem with row %s" % xrow)
                continue 
                
            else:
                xrvalues = rowValues(xr, startloc)
                
                qdata.append(xrvalues)
                if dkey >= 0:
                    key_column.append(xr[dkey])
                else:
                    key_column.append(xrow)
        
        nrows = len(qdata)
        
        if dkey >= 0:
            qdata = dict(zip(key_column, qdata))
        
        for attr in ['key_column', 'hdr', 'qdata']:
            setattr(sheetdb, attr, vars().get(attr, None))
            
        return sheetdb

    def read_sheet(self, sheet_name=None, sheet_index=0):
        def read_row(sheet, row):
            try:
                xr = map(self.xlCellValue, sh.row(row))
                
            except:
                #skips a row if there's a problem
                print("problem with row %s" % xrow)
                xr = []
                
            return xr
            
        sh = self.sheet(sheet_name, sheet_index)
        
        return [read_row(sh, xrow) for xrow in range(sh.nrows)]


class XLSXReader(object):
    def __init__(self, filename, hash_comments=True):
        self.book = None
        self.hash_comments = hash_comments
        
        # TODO: if we need to connect via url, try this kind of thing
        #  url = urlopen(filename)
        #  contents = url.read()
        #  stio = StringIO.StringIO(contents) 
        #  book = pyxl.reader.excel.load_workbook(filename = stio) 
        try:
            self.book = pyxl.reader.excel.load_workbook(filename=filename,
                                                        use_iterators=True)
        except:
            print("Error on %s" % filename)
            raise
                
    def xlCellValue(self, x): 
        if self.hash_comments and hasattr(x, "__iter__"):
            if x[0] == '#':
                return None
        try:
            return x.internal_value
        except:
            return x.value

    def sheet(self, sheet=None):
        '''
        Returns worksheet object by name or index, 
        
        Example:
        > wb = XLSXReader('test.xls')
        > wb.sheet('sheet2')   # returns sheet named 'sheet2'
        > wb.sheet(3)          # returns sheet with index 3
        
        Default (if no argument is pased) returns sheet index = 0 
        
        '''
        if type(sheet) == int:
            sheet_name, sheet_index = None, sheet
        else:
            sheet_name, sheet_index = sheet, 0
        
        try:
            if not sheet_name:
                sheet_name = self.book.get_sheet_names()[sheet_index]
        except:
            print("Invalid sheet: %s %s" % (sheet, type(sheet)))
            return None
        
        self.sh = self.book.get_sheet_by_name(sheet_name)
        
        return self.sh
        
    def sheet_as_db(self, sheet=None,
                    header=True, dkey=0,
                    startrow=0):
        '''
        Reads rows in a given sheet. Data is stored in 
        class member list self.qdata
        
        '''
        class sheetdb(object):
            pass
            
        self.sh = self.sheet(sheet)
        
        startloc = 1 if dkey == 0 else 0
        
        rowValues = lambda row_, loc: row_[loc:]
        hdr = None
        qdata = []
        key_column = []
        
        nrow = 0
        for row in self.sh.iter_rows():

            if nrow < startrow:
                nrow += 1
                continue

            else:
                if header and not hdr:
                    hdr = [cell.internal_value for cell in row]
                    def rowValues(row_, loc): 
                        return dict(zip(hdr[loc:],
                                        row_[loc:]))
                    
                else:
                    try:
                        xr = map(self.xlCellValue, row)
                        
                    except:
                        print("problem with row %s" % nrow)
                        continue #skips a row if there's a problem
                        
                    else:
                        xrvalues = rowValues(xr, startloc)
                        qdata.append(xrvalues)
                        if dkey >= 0:
                            key_column.append(xr[dkey])
                        else:
                            key_column.append(nrow)
        
        if dkey >= 0:
            qdata = dict(zip(key_column, qdata))
                
        for attr in ['key_column', 'hdr', 'qdata']:
            setattr(sheetdb, attr, vars().get(attr, "what"))
            
        return sheetdb         
